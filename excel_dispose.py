import os
import openpyxl
import re

class ExcelDispose:
	"""处理电子表格的类"""

	def __init__(self):
		"""初始化处理资源"""
		self.wb_path = input("\nPlease input your excel path: ")
		self.wb_path = self.wb_path.strip()
		self.wb = openpyxl.load_workbook(self.wb_path)
		try:
			self.sheet = self.wb['sheet1']
		except(KeyError):
			print("Worksheet sheet1 does not exist.")
			self.sheet = self.wb['Sheet1']

		self.titleName = "接收人姓名"
		self.titlePhone = "接收人电话"
		self.titleAddress = "接收人地址"
		self.target_list = [
		"咨询分类",
		"外呼-特殊项目外呼-行车记录仪U盘故障-【外呼】L9/L8U盘替换外呼（已接受）"
		]

	def run_dispose(self):
		"""开始处理的主循环"""
		# disposed_wb_path = self.delete_cells(self.sheet)
		# self.disposed_wb = openpyxl.load_workbook(disposed_wb_path)
		# self.disposed_sheet = self.disposed_wb['sheet1']
		self.create_columns(self.sheet)
		self.extract_cell_info(self.sheet)


	def delete_cells(self, wb_sheet):
		"""筛选并删除不需要的行"""
		for row in wb_sheet.rows:
			for cell in row:
				if cell.column == 8 and (cell.value not in self.target_list):
					print(f"delete: {cell.value}")
					wb_sheet.delete_rows(cell.row)
					break
		self.wb.save("disposed_ticket.xlsx")
		disposed_wb_path = "disposed_ticket.xlsx"
		return disposed_wb_path

	def extract_cell_info(self, wb_sheet):
		"""逐行提取单元格信息"""
		maxRow = wb_sheet.max_row
		for i in range(2, maxRow+1):
			cell_value = wb_sheet.cell(row=i, column=10).value
			# cell_value_split = cell_value.split('：')
			cell_value_split = re.split("：|:", cell_value)

			name_titlePhone = cell_value_split[1]
			phone_titleAddress = cell_value_split[2]
			address = cell_value_split[3]

			name_titlePhone = name_titlePhone.strip()
			name = re.split(", | |，|\n", name_titlePhone)[0]
			if "接收人电话" in name:
				name = name[:-5]
			else:
				pass

			phone_titleAddress = phone_titleAddress.strip()
			phone = re.split(", | |，|\n", phone_titleAddress)[0]
			if "接收人地址" in phone:
				phone = phone[:-5]
			else:
				pass

			print('------------------------------')
			print(f"第 {i} 行")
			print(f"{self.titleName}：{name}")
			print(f"{self.titlePhone}：{phone}")
			print(f"{self.titleAddress}：{address}")
			print('\n')

			wb_sheet.cell(row=i, column=11).value = name
			wb_sheet.cell(row=i, column=12).value = phone
			wb_sheet.cell(row=i, column=13).value = address

		self.wb.save("disposed_ticket.xlsx")

	def create_columns(self, wb_sheet):
		"""新建列以填入分解得到的信息"""
		wb_sheet.insert_cols(idx=11, amount=3)
		wb_sheet['K1'] = self.titleName
		wb_sheet['L1'] = self.titlePhone
		wb_sheet['M1'] = self.titleAddress

		self.wb.save("disposed_ticket.xlsx")
		# os.remove("disposed_ticket.xlsx")

	# def fill_data(self, wb_sheet, name, phone, address):
	# 	"""向单元格内填充数据"""
	# 	maxRow = wb_sheet.max_row
	# 	for i in range(2, maxRow+1):
	# 		wb_sheet.cell(row=i, column=11).value = name
	# 		wb_sheet.cell(row=i, column=12).value = phone
	# 		wb_sheet.cell(row=i, column=13).value = address
	# 	wb.save("disposed_ticket.xlsx")


if __name__ == '__main__':
	disposer = ExcelDispose()
	disposer.run_dispose()